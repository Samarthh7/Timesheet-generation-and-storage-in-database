import jpype
import asposecells
import xlsxwriter
import openpyxl


def create_timesheetOutput():
    jpype.startJVM()
    from asposecells.api import Workbook

    workbook = xlsxwriter.Workbook("timesheet_output.xlsx")
    workbook = Workbook(r"C:\Users\samar\Desktop\Timesheet_sample_input.xlsx")
    workbook.save("timesheet_output.xlsx")
    w1 = openpyxl.load_workbook("timesheet_output.xlsx")
    sheet1 = w1["Evaluation Warning"]
    w1.remove(sheet1)
    w1.save("timesheet_output.xlsx")
