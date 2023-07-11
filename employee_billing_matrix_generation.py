from openpyxl import workbook, load_workbook
import pymysql
import pandas as pd


# calculating working,billing days,billing percentage
def cal_workingDays(df, df_summary):
    wb = load_workbook("timesheet_output.xlsx")
    ws = wb["Summary"]
    print(ws)
    lst = list()
    count1 = 0
    for values in df.keys():
        count1 = count1 + 1
        count = 0
        days = ["Saturday", "Sunday"]
        df[values] = df[values][df[values].Day.isin(days) == False]
        df[values].drop([0, 1, 2], axis=0, inplace=True)
        for i in df[values].Hours.values:
            if i > 0:
                count = count + 1
        lst.append(count)
    print(lst)
    ws["F9"].value = "Working days"
    ws["G9"].value = "Billing days"
    ws["H9"].value = "Billing Amount"
    ws["I9"].value = "Billing %"
    company_holiday = df_summary.iat[3, 3]
    adjustment = df_summary.iat[3, 5]
    total_days = df_summary.iat[3, 1]
    billing_rate = df_summary.iat[4, 1]
    total_billing_days = total_days - (adjustment + company_holiday)
    per_day_billing = insert_billingAmount(df_summary)

    for row in range(10, 10 + count1):
        ws["F" + str(row)] = lst[0]
        ws["G" + str(row)] = lst[0] - (company_holiday + adjustment)  # billing days
        if ws["F" + str(row)].value > total_billing_days:
            ws["H" + str(row)] = billing_rate  # billing amount
        else:
            ws["H" + str(row)] = (
                ws["G" + str(row)].value * per_day_billing
            )  # billing amount
        if ws["F" + str(row)].value > total_billing_days:
            ws["I" + str(row)] = "100%"  # billing %

        else:
            ws["I" + str(row)] = (
                ws["F" + str(row)].value / total_billing_days
            ) * 100  # billing %

        del lst[0]

    wb.save("timesheet_output.xlsx")


def insert_billingAmount(df_summary):
    wb = load_workbook("timesheet_output.xlsx")
    ws = wb["Summary"]
    billing_rate = df_summary.iat[4, 1]
    total_days = df_summary.iat[3, 1]
    company_holiday = df_summary.iat[3, 3]
    adjustment = df_summary.iat[3, 5]
    total_billing_days = total_days - (adjustment + company_holiday)
    per_day_billing = billing_rate / total_billing_days
    ws["D6"].value = per_day_billing
    ws["H5"].value = total_billing_days
    wb.save("timesheet_output.xlsx")
    return per_day_billing
